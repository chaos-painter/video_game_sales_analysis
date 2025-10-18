from config import engine
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, Rule
import openpyxl.utils


def generate_charts():
    QUERIES = {
    "distribution_of_games_per_release_year": """
        select gplat.release_year, count(g.id) as game_count
        from game as g
        join game_publisher as gpub on g.id = gpub.game_id
        join game_platform as gplat on gpub.id = gplat.game_publisher_id
        group by gplat.release_year
        order by gplat.release_year
    """,
    "genres_by_sales": """
        select ge.genre_name, sum(rs.num_sales) as total_sales
        from genre as ge
        join game as g on ge.id = g.genre_id
        join game_publisher as gpub on g.id = gpub.game_id
        join game_platform as gplat on gpub.id = gplat.game_publisher_id
        join region_sales as rs on gplat.id = rs.game_platform_id
        group by ge.genre_name
        order by total_sales desc;
    """,
    "platforms_by_sales": """
        select pl.platform_name, sum(rs.num_sales) as total_sales
        from platform as pl
        join game_platform as gplat on pl.id = gplat.platform_id
        join region_sales as rs on gplat.id = rs.game_platform_id
        group by pl.platform_name
        order by total_sales desc;
    """,
    "platforms_by_game_count": """
        with ranked_platforms as (
            select 
                p.platform_name, 
                count(distinct g.id) as game_count,
                rank() over (order by count(distinct g.id) desc) as game_rank
                from platform as p
                join game_platform as gplat on p.id = gplat.platform_id
                join game_publisher as gpub on gplat.game_publisher_id = gpub.id
                join game as g on gpub.game_id = g.id
                group by p.platform_name
            )
        select 
            platform_name, 
            game_count
        from ranked_platforms
        where game_rank <= 15
        union all
        select 
            'Others' as platform_name, 
            sum(game_count) as game_count
        from ranked_platforms
        where game_rank > 15;
    """,
    "yearly_sales_of_puzzle_games": """
        select gplat.release_year, sum(rs.num_sales) as total_sales
        from genre as ge
        join game as g on ge.id = g.genre_id
        join game_publisher as gpub on g.id = gpub.game_id
        join game_platform as gplat on gpub.id = gplat.game_publisher_id
        join region_sales as rs on gplat.id = rs.game_platform_id
        where ge.genre_name = 'Puzzle'
        group by gplat.release_year
        order by gplat.release_year;
    """,
    "publisher_game_count_and_total_sales": """
        select p.publisher_name, count(distinct g.id) as game_count, sum(rs.num_sales) as total_sales
        from game as g
        join game_publisher as gpub on g.id = gpub.game_id
        join publisher as p on gpub.publisher_id = p.id
        join game_platform as gplat on gpub.id = gplat.game_publisher_id
        join region_sales as rs on gplat.id = rs.game_platform_id
        group by p.publisher_name
        order by p.publisher_name
    """
}
    # Pie chart
    data = pd.read_sql(QUERIES["platforms_by_game_count"], engine)

    plt.figure(figsize=(8, 8)) 
    data.set_index("platform_name")["game_count"].plot.pie(autopct="%1.1f%%")
    plt.title("Distribution of games per platform")
    plt.ylabel("Game count")
    plt.savefig("./charts/pie_platform_game_count.png")
    print(f"{len(data)} rows. Pie chart of distribution of games per platform.")


    # Bar chart
    data = pd.read_sql(QUERIES["genres_by_sales"], engine)

    plt.figure(figsize=(8, 8)) 
    data.plot.bar(x="genre_name", y="total_sales")
    plt.title("Genres compared by total sales")
    plt.ylabel("Total sales")
    plt.savefig("./charts/bar_genre_total_sale.png")
    print(f"{len(data)} rows. Bar chart compares genres by total sales.")


    # Horizontal bar chart
    data = pd.read_sql(QUERIES["genres_by_sales"], engine)

    plt.figure(figsize=(8, 8)) 
    data.plot.barh(x="genre_name", y="total_sales")
    plt.title("Genres compared by total sales")
    plt.ylabel("Total sales")
    plt.savefig("./charts/horizontal_bar_genre_total_sale.png")
    print(f"{len(data)} rows. Horizontal bar chart compares genres by total sales.")


    # Line chart
    data = pd.read_sql(QUERIES["yearly_sales_of_puzzle_games"], engine)

    plt.figure(figsize=(8, 8)) 
    data.plot.line(x="release_year", y="total_sales")
    plt.title("Sales of puzzle games over years")
    plt.ylabel("Total sales")
    plt.savefig("./charts/line_yearly_sales_of_puzzle_games.png")
    print(f"{len(data)} rows. Line chart of total sales of puzzle games over years")


    # Histogram
    data = pd.read_sql(QUERIES["distribution_of_games_per_release_year"], engine)

    plt.figure()
    data["release_year"].plot.hist(bins=10, weights=data["game_count"])
    plt.title("Distribution of game releases over time")
    plt.ylabel("Number of games")
    plt.savefig("./charts/hist_games_per_release_year.png")
    print(f"{len(data)} rows. Histogram of the distribution of games per release year.")


    # Scatter
    data = pd.read_sql(QUERIES["publisher_game_count_and_total_sales"], engine)

    data.plot.scatter(x="game_count", y="total_sales")
    plt.title("Correlation between game count and total sales of publishers")
    plt.ylabel("total_sales")
    plt.savefig("./charts/scatter_publisher_games_sales.png")
    print(f"{len(data)} rows. Scatter plot of publisher game count and total sales.")



def time_slider(query):

    data = pd.read_sql(query, engine)
    fig = px.bar(
        data, 
        x="genre_name", 
        y="total_sales",
        animation_frame="release_year",
        color="genre_name",
        title="Total Sales by Genre Over Time",
        category_orders={"release_year": sorted(data["release_year"].unique())}
    )
    fig.show()


def export_to_excel(query):
    data = pd.read_sql(query, engine)
    
    filename = "./exports/query_output.xlsx"
    
    data.to_excel(filename, index=False, sheet_name="Data")
    
    wb = load_workbook(filename)
    ws = wb.active
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    numeric_cols = data.select_dtypes(include=["number"]).columns.tolist()
    max_row = len(data) + 1
    
    for col_name in numeric_cols:
        col_id = data.columns.get_loc(col_name) 
        col_letter = openpyxl.utils.get_column_letter(col_id + 1)
        col_range = f"{col_letter}2:{col_letter}{max_row}"
        
        gradient_rule = ColorScaleRule(
            start_type="min", 
            start_color="FFAA0000",
            mid_type="percentile", 
            mid_value=50, 
            mid_color="FFFFFF00",
            end_type="max", 
            end_color="FF00AA00"
        )
        ws.conditional_formatting.add(col_range, gradient_rule)
        
    
    wb.save(filename)
    print(f"Data exported to {filename} with formatting applied. {len(data)} rows processed.")


if __name__ == "__main__":
    query = """
        select ge.genre_name, gplat.release_year, sum(rs.num_sales) as total_sales
        from genre as ge
        join game as g on ge.id = g.genre_id
        join game_publisher as gpub on g.id = gpub.game_id
        join game_platform as gplat on gpub.id = gplat.game_publisher_id
        join region_sales as rs on gplat.id = rs.game_platform_id
        group by ge.genre_name, gplat.release_year
        order by gplat.release_year desc
    """
    generate_charts()
    time_slider(query)
    export_to_excel(query)



